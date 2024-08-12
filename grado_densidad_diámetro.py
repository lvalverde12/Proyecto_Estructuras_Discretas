import csv
import networkx as nx
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment

def load_graph_from_csv(csv_file):
    G = nx.Graph()
    with open(csv_file, newline='', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        for row in reader:
            ruta = row['Código CTP']
            inicio = row['Código Distrito Inicio']
            fin = row['Código Distrito Final']
            extension = float(row['Extensión (km)'])
            G.add_node(inicio, label=row['Cantón Inicio'])
            G.add_node(fin, label=row['Cantón Final'])
            G.add_edge(inicio, fin, weight=extension, label=ruta)
    return G

def analyze_graph(G):
    degrees = dict(G.degree())
    density = nx.density(G)
    try:
        diameter = nx.diameter(G)
    except nx.NetworkXError:
        diameter = "El grafo no es conexo, por lo que no tiene un diámetro definido."
    return degrees, density, diameter

def export_to_xlsx(degrees, density, diameter, filename):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Grado de los nodos"
    ws['A1'].alignment = Alignment(horizontal='center')
    for i, (node, degree) in enumerate(degrees.items(), start=2):
        ws[f'A{i}'] = f"Nodo {node}"
        ws[f'A{i}'].alignment = Alignment(horizontal='center')
        ws[f'B{i}'] = degree
        ws[f'B{i}'].alignment = Alignment(horizontal='center')

    ws[f'A{i+2}'] = "Densidad del grafo"
    ws[f'A{i+2}'].alignment = Alignment(horizontal='center')
    ws[f'B{i+2}'] = density
    ws[f'B{i+2}'].number_format = '0.0000'  
    ws[f'B{i+2}'].alignment = Alignment(horizontal='center')

    ws[f'A{i+3}'] = "Diámetro del grafo"
    ws[f'A{i+3}'].alignment = Alignment(horizontal='center')
    ws[f'B{i+3}'] = diameter
    ws[f'B{i+3}'].alignment = Alignment(horizontal='center')
    wb.save(filename)

csv_file = "Rutas.csv"
G = load_graph_from_csv(csv_file)

degrees, density, diameter = analyze_graph(G)

export_to_xlsx(degrees, density, diameter, "Grado_Densidad_Diámetro_analysis.xlsx")

print("Grado de los nodos:")
for node, degree in degrees.items():
    print(f"Nodo {node}: {degree}")

print(f"\nDensidad del grafo: {density:.4f}")
print(f"Diámetro del grafo: {diameter}")
