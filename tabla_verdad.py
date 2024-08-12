import csv
import networkx as nx
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment

def load_graph_from_csv(csv_file):
    G = nx.Graph()
    with open(csv_file, newline='', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        for row in reader:
            inicio = row['Código Distrito Inicio']
            fin = row['Código Distrito Final']
            G.add_edge(inicio, fin)
    return G

def save_truth_table_to_excel(G, excel_file):
    nodes = list(G.nodes)
    matrix = nx.to_numpy_array(G, nodelist=nodes)
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabla de Verdad"

    for i, node in enumerate(nodes, start=2):
        ws.cell(row=1, column=i, value=node).alignment = Alignment(horizontal='center')
        ws.cell(row=i, column=1, value=node).alignment = Alignment(horizontal='center')

    for i, row in enumerate(matrix, start=2):
        for j, value in enumerate(row, start=2):
            cell = ws.cell(row=i, column=j, value=int(value))
            cell.number_format = '0' 
            cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(excel_file)

def main():
    csv_file = 'Rutas.csv'
    excel_file = 'Tabla_Verdad.xlsx'
    G = load_graph_from_csv(csv_file)
    save_truth_table_to_excel(G, excel_file)

if __name__ == '__main__':
    main()
